#importing openpyxl and as means instead of calling openpyxl access by just xl
import openpyxl as xl
#loading the workbook and it returns the workbook object wb
wb=xl.load_workbook('transactions.xlsx')


#from a number of sheets we can access the particular sheet
sheet=wb['Sheet1']
#aproaches to access the cell
cell=sheet['a1']
#OR
# cell=sheet.cell(1,1)
print(cell.value)
#sheet.max row +1 for loop desnt take the upto max row and nit take theat particular value
for row in range(2,sheet.max_row+1):
    cell=sheet.cell(row,3)#(row,colounn)
    corrected_price=cell.value*0.9
    corrected_price_cell=sheet.cell(row,5)
    corrected_price_cell.value=corrected_price

k=sheet.cell(1,5)
k.value="corrected price"
wb.save("transaction2.xlsx")

